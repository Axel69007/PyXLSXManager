import openpyxl 
from openpyxl import *
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
from tkinter import simpledialog
import tkinter as tk
from tkinter import PhotoImage 

#Déclaration variable global
#Tkinter
root = tk.Tk()
root.title("PyXLSXManager")
root.minsize(300, 300)
#Excel
wb = load_workbook('hello.xlsx')
ws = wb.active

#Déclaration Fonction
def quitter_programme():
    root.destroy()

def afficher_data():
	global wb,ws,root
	max_row_minus_ten = ws.max_row - 10
	max_row = ws.max_row
	top_fenetre=Toplevel(root)
	top_fenetre.minsize(1000, 250)
	my_scrollbar = Scrollbar(top_fenetre, orient=VERTICAL)
	my_listbox = Listbox(top_fenetre, width=100, yscrollcommand=my_scrollbar.set, selectmode=NONE)
	my_listbox.config
	my_scrollbar.pack(side=RIGHT, fill=Y)
	my_listbox.pack(pady=15, fill=X)
	#Lecture fichier et sauvegarde des 10 dernière lignes
	my_list = []
	for value in ws.iter_rows(min_row=max_row_minus_ten, max_row=max_row, min_col=1, max_col=8,values_only=True):
		my_list.append(str(value))
	#Insertion dans la lisbox
	for item in my_list:
		my_listbox.insert(END, item)
	

def ajouter_data():
	global wb,ws

	max_row_plus_one = ws.max_row + 1
	
	top_fenetre=Toplevel(root)
	top_fenetre.minsize(500, 250)
	top_fenetre.columnconfigure((0,8),weight =1,uniform='a')
	top_fenetre.columnconfigure((1,2,3,4,5,6,7),weight =2,uniform='a')
	top_fenetre.rowconfigure((0,1,2,3),weight =1,uniform='a')

	label_information = Label(master= top_fenetre, text="Veuillez ajouter les données ci dessous : ", font="Helvetica 18 bold")
	label_information.grid(row = 0,column = 0, sticky="n",columnspan=10)

	label_ref = Label(master= top_fenetre, text="REF : ", font="Helvetica 14 ")
	label_ref.grid(row = 1,column = 1,sticky="nwse")
	entry_ref = tk.Entry(top_fenetre, font=("Helvetica", 12))
	entry_ref.grid(row = 2,column = 1,sticky="we")

	label_date_arrive = Label(master= top_fenetre, text="Date arrivée : ", font="Helvetica 14")
	label_date_arrive.grid(row = 1,column = 2,sticky="nwse")
	entry_date_arrive = tk.Entry(top_fenetre, font=("Helvetica", 12))
	entry_date_arrive.grid(row = 2,column = 2,sticky="we")

	label_date_fin = Label(master= top_fenetre, text="Date fin : ", font="Helvetica 14")
	label_date_fin.grid(row = 1,column = 3,sticky="nwse")
	entry_date_fin = tk.Entry(top_fenetre, font=("Helvetica", 12))
	entry_date_fin.grid(row = 2,column = 3,sticky="we")

	label_status = Label(master= top_fenetre, text="Status : ", font="Helvetica 14")
	label_status.grid(row = 1,column = 4,sticky="nwse")
	entry_status = tk.Entry(top_fenetre, font=("Helvetica", 12))
	entry_status.grid(row = 2,column = 4,sticky="we")

	label_responsable = Label(master= top_fenetre, text="Responsable : ", font="Helvetica 14")
	label_responsable.grid(row = 1,column = 5,sticky="nwse")
	entry_responsable = tk.Entry(top_fenetre, font=("Helvetica", 12))
	entry_responsable.grid(row = 2,column = 5,sticky="we")

	label_description = Label(master= top_fenetre, text="Description : ", font="Helvetica 14")
	label_description.grid(row = 1,column = 6,sticky="nwse")
	entry_description = tk.Entry(top_fenetre, font=("Helvetica", 12))
	entry_description.grid(row = 2,column = 6,sticky="we")

	label_technologie = Label(master= top_fenetre, text="Technologie : ", font="Helvetica 14")
	label_technologie.grid(row = 1,column = 7,sticky="nwse")
	entry_technologie = tk.Entry(top_fenetre, font=("Helvetica", 12))
	entry_technologie.grid(row = 2,column = 7,sticky="we")

	

	def ecriture_excel():
		global wb,ws
		numero_ref=entry_ref.get()
		date_arrive=entry_date_arrive.get()
		date_fin=entry_date_fin.get()
		status=entry_status.get()
		responsable=entry_responsable.get()
		description=entry_description.get()
		technologie=entry_technologie.get()


		ws.cell(row=max_row_plus_one,column=1,value=numero_ref)
		ws.cell(row=max_row_plus_one,column=2,value=date_arrive)
		ws.cell(row=max_row_plus_one,column=3,value=date_fin)
		ws.cell(row=max_row_plus_one,column=4,value=status)
		ws.cell(row=max_row_plus_one,column=5,value=responsable)
		ws.cell(row=max_row_plus_one,column=6,value=description)
		ws.cell(row=max_row_plus_one,column=7,value=technologie)

		wb.save("hello.xlsx")


		top_fenetre.destroy()

	bouton_sauvegarder = Button(top_fenetre, text="Sauvegarder",font="Helvetica 12" ,command=ecriture_excel)
	bouton_sauvegarder.grid(row = 3,column = 4, sticky="n")

def modifier_data():
	global wb,ws,root
	max_row_minus_ten = ws.max_row - 10
	max_row = ws.max_row
	top_fenetre=Toplevel(root)
	top_fenetre.minsize(1000, 250)
	my_scrollbar = Scrollbar(top_fenetre, orient=VERTICAL)
	my_listbox = Listbox(top_fenetre, width=100, yscrollcommand=my_scrollbar.set, selectmode=SINGLE)
	my_listbox.config
	my_scrollbar.pack(side=RIGHT, fill=Y)
	my_listbox.pack(pady=15, fill=X)
	#Lecture fichier et sauvegarde des 10 dernière lignes
	my_list = []
	for value in ws.iter_rows(min_row=max_row_minus_ten, max_row=max_row, min_col=1, max_col=8,values_only=True):
		my_list.append(str(value))
	#Insertion dans la lisbox
	for item in my_list:
		my_listbox.insert(END, item)
	def modification_fichier():
		#Sélection user
		selection_user = list(my_listbox.curselection())
		index_selection = 10 - selection_user[0]
		user_selection_row= ws.max_row - index_selection
		#Ancienne valeur pour préremplissage
		old_ref=ws.cell(row=user_selection_row, column=1).value
		old_date_arrive=ws.cell(row=user_selection_row, column=2).value
		old_date_fin=ws.cell(row=user_selection_row, column=3).value
		old_status=ws.cell(row=user_selection_row, column=4).value
		old_responsable=ws.cell(row=user_selection_row, column=5).value
		old_description=ws.cell(row=user_selection_row, column=6).value
		old_technologie=ws.cell(row=user_selection_row, column=7).value
		#Config toplevel
		top_fenetre_1=Toplevel(root)

		top_fenetre_1.minsize(500, 250)
		top_fenetre_1.columnconfigure((0,8),weight =1,uniform='a')
		top_fenetre_1.columnconfigure((1,2,3,4,5,6,7),weight =2,uniform='a')
		top_fenetre_1.rowconfigure((0,1,2,3),weight =1,uniform='a')

		label_information = Label(master= top_fenetre_1, text="Veuillez ajouter les données ci dessous : ", font="Helvetica 18 bold")
		label_information.grid(row = 0,column = 0, sticky="n",columnspan=10)

		label_ref = Label(master= top_fenetre_1, text="REF : ", font="Helvetica 14 ")
		label_ref.grid(row = 1,column = 1,sticky="nwse")
		entry_ref = tk.Entry(top_fenetre_1, font=("Helvetica", 12))
		entry_ref.insert(0, old_ref)
		entry_ref.grid(row = 2,column = 1,sticky="we")

		label_date_arrive = Label(master= top_fenetre_1, text="Date arrivée : ", font="Helvetica 14")
		label_date_arrive.grid(row = 1,column = 2,sticky="nwse")
		entry_date_arrive = tk.Entry(top_fenetre_1, font=("Helvetica", 12))
		entry_date_arrive.insert(0, old_date_arrive)
		entry_date_arrive.grid(row = 2,column = 2,sticky="we")

		label_date_fin = Label(master= top_fenetre_1, text="Date fin : ", font="Helvetica 14")
		label_date_fin.grid(row = 1,column = 3,sticky="nwse")
		entry_date_fin = tk.Entry(top_fenetre_1, font=("Helvetica", 12))
		entry_date_fin.insert(0, old_date_fin)
		entry_date_fin.grid(row = 2,column = 3,sticky="we")

		label_status = Label(master= top_fenetre_1, text="Status : ", font="Helvetica 14")
		label_status.grid(row = 1,column = 4,sticky="nwse")
		entry_status = tk.Entry(top_fenetre_1, font=("Helvetica", 12))
		entry_status.insert(0,old_status)
		entry_status.grid(row = 2,column = 4,sticky="we")

		label_responsable = Label(master= top_fenetre_1, text="Responsable : ", font="Helvetica 14")
		label_responsable.grid(row = 1,column = 5,sticky="nwse")
		entry_responsable = tk.Entry(top_fenetre_1, font=("Helvetica", 12))
		entry_responsable.insert(0,old_responsable)
		entry_responsable.grid(row = 2,column = 5,sticky="we")

		label_description = Label(master= top_fenetre_1, text="Description : ", font="Helvetica 14")
		label_description.grid(row = 1,column = 6,sticky="nwse")
		entry_description = tk.Entry(top_fenetre_1, font=("Helvetica", 12))
		entry_description.insert(0,old_description)
		entry_description.grid(row = 2,column = 6,sticky="we")

		label_technologie = Label(master= top_fenetre_1, text="Technologie : ", font="Helvetica 14")
		label_technologie.grid(row = 1,column = 7,sticky="nwse")
		entry_technologie = tk.Entry(top_fenetre_1, font=("Helvetica", 12))
		entry_technologie.insert(0,old_technologie)
		entry_technologie.grid(row = 2,column = 7,sticky="we")

		def modification_excel():
			global wb,ws
			numero_ref=entry_ref.get()
			date_arrive=entry_date_arrive.get()
			date_fin=entry_date_fin.get()
			status=entry_status.get()
			responsable=entry_responsable.get()
			description=entry_description.get()
			technologie=entry_technologie.get()


			ws.cell(row=user_selection_row,column=1,value=numero_ref)
			ws.cell(row=user_selection_row,column=2,value=date_arrive)
			ws.cell(row=user_selection_row,column=3,value=date_fin)
			ws.cell(row=user_selection_row,column=4,value=status)
			ws.cell(row=user_selection_row,column=5,value=responsable)
			ws.cell(row=user_selection_row,column=6,value=description)
			ws.cell(row=user_selection_row,column=7,value=technologie)

			wb.save("hello.xlsx")


			top_fenetre.destroy()
			top_fenetre_1.destroy()

		bouton_modifier= Button(master= top_fenetre_1, text="Modifier", command=modification_excel, font="Helvetica 12")
		bouton_modifier.grid(row=3,column=4)

	bouton_modifier= Button(top_fenetre, text="Modifier", command=modification_fichier, font="Helvetica 12")
	bouton_modifier.pack(side=BOTTOM,pady=10)

def supprimer_data():
	global wb,ws,root
	max_row_minus_ten = ws.max_row - 10
	max_row = ws.max_row
	top_fenetre=Toplevel(root)
	top_fenetre.minsize(1000, 250)
	my_scrollbar = Scrollbar(top_fenetre, orient=VERTICAL)
	my_listbox = Listbox(top_fenetre, width=100, yscrollcommand=my_scrollbar.set, selectmode=SINGLE)
	my_listbox.config
	my_scrollbar.pack(side=RIGHT, fill=Y)
	my_listbox.pack(pady=15, fill=X)
	#Lecture fichier et sauvegarde des 10 dernière lignes

	my_list = []
	for value in ws.iter_rows(min_row=max_row_minus_ten, max_row=max_row, min_col=1, max_col=8,values_only=True):
		my_list.append(str(value))
	#Insertion dans la lisbox
	for item in my_list:
		my_listbox.insert(END, item)
	
	
	def suppression_excel():
		global wb,ws
		selection_user = list(my_listbox.curselection())
		index_selection = 10 - selection_user[0]
		user_selection_row= ws.max_row - index_selection
		
		ws.delete_rows(user_selection_row)

		wb.save("hello.xlsx")

		my_listbox.delete(0, END)  # Vider la listbox
		max_row_minus_ten = ws.max_row - 10
		max_row = ws.max_row
		for value in ws.iter_rows(min_row=max(1, max_row_minus_ten), max_row=max_row, min_col=1, max_col=8, values_only=True):
			my_listbox.insert(END, str(value))

	def quitter_toplevel():
		top_fenetre.destroy()

	bouton_quitter_1= Button(top_fenetre, text="Quitter", command=quitter_toplevel, font="Helvetica 12")
	bouton_quitter_1.pack(side=BOTTOM,pady=10)
	bouton_supprimer_1= Button(top_fenetre, text="Supprimer", command=suppression_excel, font="Helvetica 12")
	bouton_supprimer_1.pack(side=BOTTOM,pady=10)
	



#Grid
root.columnconfigure((0),weight =1,uniform='a')
root.rowconfigure((0,1,2,3,4,5,6,7),weight =1,uniform='a')

label_bienvenue = Label(master= root, text="Bienvenue dans PyXLSXManager ", font="Helvetica 18 bold")
label_bienvenue.grid(row = 0,column = 0, sticky="n")

image_main = PhotoImage(file="png/main_image.png")
label_image = Label(master = root, image= image_main)
label_image.grid(row = 1,column = 0,sticky="nwse")

bouton_ajouter = Button(master= root, text="Ajoutez les informations", font="Helvetica 12",command=ajouter_data , width = 25)
bouton_ajouter.grid(row = 5,column = 0, sticky="n")

bouton_modifier = Button(master= root, text="Modifier les dernières entrées",font="Helvetica 12",command=modifier_data , width = 25)
bouton_modifier.grid(row = 4,column = 0, sticky="n")

bouton_voir = Button(master= root, text="Voir les dernières entrées",font="Helvetica 12",command=afficher_data , width = 25)
bouton_voir.grid(row = 3,column = 0, sticky="n")

bouton_supprimer = Button(master= root, text="Supprimer les dernières entrées",font="Helvetica 12",command=supprimer_data , width = 25)
bouton_supprimer.grid(row = 6,column = 0, sticky="n")

bouton_quitter = Button(root, text="Quitter",font="Helvetica 12" ,command=quitter_programme)
bouton_quitter.grid(row = 7,column = 0, sticky="n")



# Save the file
wb.save("hello.xlsx")
#Mainloop
root.mainloop()

#To do :
#Préremblisage des champs pour la modification
#Affichage des ref du moi
#Affichage en fonction de l'état
#Ajout vérification des champs
